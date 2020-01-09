from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from itertools import islice
from random import randint


def get_results():
    results = []

    with open("results_v1.txt", "r") as fi:
        for entry in islice(fi, 2, None):
            result = entry.strip()
            results.append(result)

    return results


def excel_export():
    results = get_results()[-100:]

    wb = Workbook()
    sheet = wb.active

    header_gap = 3
    header_start = 1
    header_end = 3

    header_font = Font(size=15, bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="33ccff")
    alignment = Alignment(horizontal="center", vertical="center")

    digit_font = Font(size=15)

    for num in range(0, 10):
        sheet.merge_cells(start_row=1, end_row=1,
                          start_column=header_start,
                          end_column=header_end)

        selectedCell = sheet.cell(1, header_start)
        selectedCell.font = header_font
        selectedCell.alignment = alignment
        selectedCell.fill = header_fill
        selectedCell.value = num

        digit_fill = PatternFill(
            fill_type="solid", fgColor=str(randint(696969, 778899)))

        for rowCount, result in enumerate(results, start=2):
            for colCount, digit in enumerate(result):
                sheet.cell(
                    rowCount, colCount + header_start).font = digit_font
                sheet.cell(
                    rowCount, colCount + header_start).alignment = alignment
                sheet.cell(
                    rowCount, colCount + header_start).value = int(digit)

                if num == int(digit):
                    sheet.cell(
                        rowCount, colCount + header_start).fill = digit_fill

        header_start = header_end + 2
        header_end = header_end + header_gap + 1

    wb.save("digit_position.xlsx")


def main():

    print("Generating excel file, please wait...")

    excel_export()

    print("Done exporting.")


if __name__ == "__main__":
    main()
