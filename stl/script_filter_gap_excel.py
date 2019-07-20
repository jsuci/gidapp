from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from itertools import islice
from random import randint
from pathlib import Path


def get_all_results():
    """
    INPUT:
        results_v1.1.txt - a text file containing all the results in vertical
        format

    OUTPUT:
        output - a list of results (ex. ["123", "456"...]) taken from
        results_v1.1.txt
    """

    output = []

    with open("results_v1.txt", "r") as fi:
        for entry in islice(fi, 2, None):
            output.insert(0, entry.strip())

    return output


def get_gap_results(num, step):
    """
    INPUT:
        all_results - a list of results (ex. ["427", "691", ...]) taken from
        get_all_results()

    OUTPUT:
        output - a dictionary composed of numbers(0-9) as keys and values
        consists of list of tuples containing the gap value and marked results
        (ex. {0: [(1, ["123", ("456", True)])]})

    """

    all_results = get_all_results()
    gap = step
    found_count = 0
    output = {}
    output.setdefault(num, [step, []])

    for count, result in enumerate(all_results):
        if step == count:
            if num in result:
                # print(f"{result}*")
                output[num][1].insert(0, (result, True))

                step += (gap + 1)
                found_count += 1

        else:
            # print(f"{result}")
            output[num][1].insert(0, (result, False))

    if found_count >= 3:
        return output

    else:
        return None


def remove_existing_files():
    """
    INPUT:
        excel_files - a list of excel_files in pathlib format (ex.
        [PosixPath('pathlib.py'), PosixPath('setup.py'))
    """

    excel_files = Path(".").glob("results_filter_gap_excel_*.xlsx")

    for excel_file in excel_files:
        print(f"removing {excel_file}")
        excel_file.unlink()


def filter_gap_excel():
    """
    INPUT:
        num - a string number(ex. "0") from 0 to 9
        step - an int(ex. 1) from 1 to n number of gap

    OUTPUT:
        filter_gap_excel_0_1.xlxs - an excel file that contains
        marked digit with gap
    """

    remove_existing_files()

    for num in range(0, 10):

        # Control the nuber of gaps here
        for step in range(1, 20):
            all_gap_results = get_gap_results(str(num), step)

            if all_gap_results:
                for k, v in all_gap_results.items():
                    gap_num = k
                    gap_step = v[0]

                    # Control the amount of results
                    gap_res = v[1][-100:]

                    # Print results found
                    print(f"{gap_num}, {gap_step}, {str(gap_res):.10}...")

                    # Excel styling options
                    digit_align = Alignment(
                        horizontal="center", vertical="center")
                    digit_font = Font(size=15)
                    digit_fill = PatternFill(
                        fill_type="solid",
                        fgColor=str(randint(696969, 778899))
                    )

                    # Create excel workbook
                    wb = Workbook()
                    sheet = wb.active

                    for row_count, res_bool in enumerate(gap_res, start=1):
                        result = res_bool[0]
                        match = res_bool[1]
                        # found_digits = []

                        for col_count, digit in enumerate(result, start=1):

                            # Styling each cell and filtering matches
                            sheet.cell(row_count, col_count).font = digit_font

                            sheet.cell(row_count,
                                       col_count
                                       ).alignment = digit_align

                            sheet.cell(
                                row_count,
                                col_count
                            ).value = int(digit)

                            if (
                                match and
                                gap_num == digit
                                # digit not in found_digits
                            ):
                                # found_digits.append(digit)
                                sheet.cell(
                                    row_count,
                                    col_count
                                ).fill = digit_fill

                    wb.save(
                        f"results_filter_gap_excel_{gap_num}_{gap_step}.xlsx")


def main():
    filter_gap_excel()


if __name__ == "__main__":
    main()
