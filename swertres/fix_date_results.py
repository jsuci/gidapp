from itertools import islice, chain
from re import split
from openpyxl import Workbook
from openpyxl.styles import (NamedStyle, Alignment, Font)
from datetime import datetime


def read_results():
    output = []

    with open("results_v2.txt", "r") as fi:
        # updated: 04 sat jan 2020 2
        check_date = fi.readline().strip()[-1]

        for entry in islice(fi, 1, None):
            entry = split(r"\s{2,}", entry.strip())
            entry[0] = datetime.strptime(
                entry[0],
                "%d %a %b %Y"
            ).strftime(
                "%d %B %Y"
            )

            output.append(entry)

        if check_date == "2":
            return output
        else:
            return output[:-1]


def fix_results():
    all_results = read_results()
    month_results = []
    temp_results = []
    output = []

    for i in range(1, len(all_results)):
        # pev_month = January
        prev_month = all_results[i - 1][0].split(" ")[1]
        next_month = all_results[i][0].split(" ")[1]

        if prev_month == next_month:
            temp_results.append(all_results[i - 1])
        else:
            temp_results.append(all_results[i - 1])
            month_results.append(temp_results)
            temp_results = []

    # Append the last entry
    temp_results.append(all_results[-1])
    month_results.append(temp_results)

    for month_res in month_results:
        for i in range(32 - 1):
            date_one = int(f"{i + 1:02d}")
            add_entry = [
                f"{date_one:02d} {month_res[0][0][3:]}",
                "",
                "",
                ""
            ]

            try:
                month_res_index = month_res.index(month_res[i])
                date_two = int(month_res[i][0][:2])
            except Exception:
                month_res.append(add_entry)
                continue

            if date_one != date_two:
                date_one_int = i
                date_two_int = int(month_res[i][0][:2])

                if date_one_int < date_two_int:
                    month_res.insert(month_res_index, add_entry)

        output.append(month_res)

    for e in list(chain(*output)):
        print(e)


def excel_export():
    results = fix_results()

    # Creating Workbook
    wb = Workbook()
    ws = wb.active

    # Formatting
    cell_style = NamedStyle(
        name="cell_style",
        font=Font(bold=True),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    wb.add_named_style(cell_style)

    # Control group
    month_gap = 1
    year_gap = 1
    time_str = ["11AM", "4PM", "9PM"]

    for count, res in enumerate(results, start=1):
        if (res == "end_month"):
            month_gap += 3
        elif (res == "end_year"):
            year_gap += 35
            month_gap = 1
        else:

            # Date header
            ws.merge_cells(
                start_row=year_gap,
                end_row=year_gap + 1,
                start_column=1,
                end_column=1
            )

            date_cell = ws.cell(row=year_gap, column=1, value="Date")
            date_cell.style = cell_style

            # ['02', 'January', '2013']
            res_date = res[0].split(" ")
            res_day = res_date[0]
            res_month = res_date[1]
            res_year = res_date[2]

            digits = res[1:]

            # Month header
            ws.merge_cells(
                start_row=year_gap,
                end_row=year_gap,
                start_column=month_gap + 1,
                end_column=month_gap + 3
            )

            month_cell = ws.cell(
                row=year_gap,
                column=month_gap + 1,
                value=f"{res_month} {res_year}")
            month_cell.style = cell_style

            # # Time header
            for c_time, time in enumerate(time_str, start=1):
                ws.cell(
                    row=year_gap + 1,
                    column=month_gap + c_time,
                    value=time
                ).style = cell_style

            # Date entry
            ws.cell(
                row=year_gap + count + 1,
                column=1,
                value=res_day
            ).style = cell_style

            # # Result entry
            # for dg_c, dg in enumerate(digits):
            #     ws.cell(
            #         row=count + 2,
            #         column=dg_c + 2,
            #         value=dg
            #     ).style = cell_style

    wb.save("fix_date_results.xlsx")


def main():
    # excel_export()
    fix_results()


if __name__ == "__main__":
    main()
