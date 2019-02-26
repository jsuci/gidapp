# from itertools import *


# def get_results():
#     """Get all the previous result and store it in reverse order"""

#     results = []
#     with open("results_v1.txt", "r") as fi:
#         for line in islice(fi, 2, None):
#             results.insert(0, line.strip())

#     return results


# def diff_one(digit):
#     return {
#         0: 1,
#         1: 2,
#         2: 3,
#         3: 4,
#         4: 5,
#         5: 6,
#         6: 7,
#         7: 8,
#         8: 9,
#         9: 0
#     }[digit]


# def diff_two(digit):
#     return {
#         0: 2,
#         1: 3,
#         2: 4,
#         3: 5,
#         4: 6,
#         5: 7,
#         6: 8,
#         7: 9,
#         8: 0,
#         9: 1
#     }[digit]


# def seq_type(digits):
#     """
#     Given a sequence of string numbers ['8', '9', '0', '1'...] determine
#     what type of sequence it has. Return a string of seq_type
#         diff_one - if all numbers has a difference of 1
#         diff_two - if all numbers has a difference of 2
#         diff_zero - if all numbers has the same digit

#         gap_one - all of the numbers has a difference of one except 1
#             ex. 8, 9, 0, 1, 3 (2 missing)
#                 7, 9, 0, 1, 2 (8 missing)
#     """

#     uniq_digits = set([int(e) for e in digits])
#     diff_one_count = 0
#     diff_two_count = 0
#     diff_none_count = 0

#     if len(uniq_digits) == 1:
#         return "diff_zero"
#     elif len(uniq_digits) != len(digits):
#         return "has_double"
#     else:
#         for digit in uniq_digits:

#             if diff_one(digit) in uniq_digits:
#                 diff_one_count += 1
#             elif diff_two(digit) in uniq_digits:
#                 diff_two_count += 1
#             else:
#                 diff_none_count += 1

#         # print(diff_one_count, diff_two_count, diff_none_count)

#         # Filter diff_one
#         if diff_two_count == 0 and diff_none_count < 2:
#             return "diff_one"

#         # Filter diff_two
#         if (
#             diff_one_count == 0 and
#             diff_two_count >= 1 and
#             diff_none_count != 2
#         ):
#             return "diff_two"

#         # Filter gap_one
#         if (
#             diff_one_count != 0 and
#             diff_two_count == 1 and
#             diff_none_count != 2
#         ):
#             return "gap_one"


# def plus_one(digit):
#     digit = int(digit)

#     return {
#         0: 1,
#         1: 2,
#         2: 3,
#         3: 4,
#         4: 5,
#         5: 6,
#         6: 7,
#         7: 8,
#         8: 9,
#         9: 0
#     }[digit]


# def minus_one(digit):
#     digit = int(digit)

#     return {
#         1: 0,
#         2: 1,
#         3: 2,
#         4: 3,
#         5: 4,
#         6: 5,
#         7: 6,
#         8: 7,
#         9: 8,
#         0: 9
#     }[digit]


# def diff_one_pair(diff_one_digits):
#     """Given a list of diff_one_digits(['1', '2', '3'..])
#     determine the possible next digit. Return a list of
#     next digit. ex: [0, 4] from ['1', '2', '4']
#     """

#     all_pairs = []

#     sort_entry = sorted(diff_one_digits)

#     if (
#         '0' in sort_entry and
#         '1' in sort_entry and
#         '9' in sort_entry
#     ):
#         start_digit = minus_one('9')
#         last_digit = plus_one('1')

#     elif (
#         '0' in sort_entry and
#         '8' in sort_entry and
#         '9' in sort_entry
#     ):
#         start_digit = minus_one('8')
#         last_digit = plus_one('0')

#     else:
#         start_digit = minus_one(sort_entry[0])
#         last_digit = plus_one(sort_entry[-1])

#     all_pairs = [start_digit, last_digit]

#     return all_pairs


# def filter_results():
#     """Process get_results() output and filter them by gap.
#     Return a list of tuple containing [(gap_value, results)]
#     """

#     results = get_results()
#     final_list = []

#     for gap_value in range(1, 100):
#         step = gap_value
#         temp_results = []
#         temp_digits = []

#         for count, result in enumerate(results):
#             if step == count and len(temp_results) != 3:
#                 temp_results.append(result)
#                 temp_digits.append(result[0])
#                 step += (gap_value + 1)

#         pair = diff_one_pair(temp_digits)
#         final_list.append((
#             gap_value, temp_results, temp_digits, pair))

#     return final_list


# def main():
#     p_first_digits = []

#     for entry in filter_results():
#         gap, results, digits, pair = entry

#         if seq_type(digits) == "diff_one":
#             p_first_digits.extend(pair)
#             print("gap: {}".format(gap))
#             print("results: {}".format(results))
#             print("digits: {}".format(digits))
#             print("possible_digits: {}".format(pair))
#             print("\n")

#     print(set(p_first_digits))
#     print("\n")


# if __name__ == "__main__":
#     main()

##############################
# from itertools import *


# def all_equal(digit_1, digit_2):
#     """Given digit_1('123') and digit_2 compare if both
#     are equal. Return True if they are and False if not
#     """

#     for num_1 in digit_1:
#         if num_1 in digit_2:
#             digit_2 = digit_2.replace(num_1, "", 1)

#     # Check if all 3 digits matched
#     if len(digit_2) == 0:
#         return True
#     else:
#         return False


# def two_equal(digit_1, digit_2):
#     """Given digit_1('123') and digit_2 compare if both
#     are equal. Return True if they are and False if not
#     """

#     for num_1 in digit_1:
#         if num_1 in digit_2:
#             digit_2 = digit_2.replace(num_1, "", 1)

#     # Check if all 3 digits matched
#     if len(digit_2) <= 1:
#         return True
#     else:
#         return False


# def search_results(number, pair):
#     """Given a string number, find all occurance of
#      that number and get the top and bottom results.
#     Return a list of tuple containing the [(top, mid, bot)]
#     """

#     with open("results_v1.txt", "r") as fi:
#         results = [e.strip() for e in fi]
#         i = 2
#         common_res = []

#         while i < len(results) - 1:
#             top = results[i - 1]
#             mid = results[i]
#             bot = results[i + 1]

#             if all_equal(number, mid):
#                 print(top, mid, bot)

#                 sorted_top = "".join(sorted(top))
#                 sorted_bot = "".join(sorted(bot))

#                 if (two_equal(pair, top)):
#                     common_res.append(sorted_top)

#                 if (two_equal(pair, bot)):
#                     common_res.append(sorted_bot)

#             i += 1

#         print(sorted(common_res))


# def main():
#     search_results("332", "12")


# if __name__ == "__main__":
#     main()

# ######################################
# # from operator import itemgetter

# # a = [[3, 'hello', 'Hello'], [2, 'cat', 'dog'], [6, 'mini', 'mouse']]

# # b = list(map(lambda x: x[0], a))

# # print(b)

# #####################################
# from itertools import *


# def get_reverse_result():
#     results = []
#     with open("results_v1.txt", "r") as fi:
#         for entry in islice(fi, 2, None):
#             result = entry.strip()
#             results.insert(0, result)

#     return results


# def count_missing_digit(digit):
#     """Given a str digit determine the longest missing
#     position of that digit. Return a list containing
#     digit, {position: missing_count}, "0 _ _")
#     """

#     results = get_reverse_result()
#     first_count = 0
#     second_count = 0
#     third_count = 0
#     highest_count = 0
#     highest_format = ""

#     final_result = [digit, {"first": 0, "second": 0, "third": 0}]

#     for result in results:
#         if digit != result[0]:
#             first_count += 1
#         else:
#             final_result[1]["first"] = first_count
#             if highest_count < first_count:
#                 highest_count = first_count
#                 highest_format = "{} - -".format(digit)
#             break

#     for result in results:
#         if digit != result[1]:
#             second_count += 1
#         else:
#             final_result[1]["second"] = second_count
#             if highest_count < second_count:
#                 highest_count = second_count
#                 highest_format = "- {} -".format(digit)
#             break

#     for result in results:
#         if digit != result[2]:
#             third_count += 1
#         else:
#             final_result[1]["third"] = third_count
#             if highest_count < third_count:
#                 highest_count = third_count
#                 highest_format = "- - {}".format(digit)
#             break

#     final_result.extend([highest_count, highest_format])
#     return final_result


# def all_missing_digit():
#     for i in range(10):
#         print(count_missing_digit(str(i))[0], "<-",
#               count_missing_digit(str(i))[2], " ",
#               count_missing_digit(str(i))[3])


# def main():
#     all_missing_digit()


# if __name__ == "__main__":
#     main()
####################################
"""
Get all results from results_v1.txt and transfer it
to excel with some digits being filtered
"""

from openpyxl import *
from openpyxl.utils import *
from openpyxl.styles import *
from itertools import *
from random import *


def get_results():
    results = []

    with open("results_v1.txt", "r") as fi:
        for entry in islice(fi, 2, None):
            result = entry.strip()
            results.append(result)

    return results


def export_to_excel():
    results = get_results()

    wb = Workbook()
    sheet = wb.active

    header_gap = 3
    header_start = 1
    header_end = 3

    header_font = Font(size=15, bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="33ccff")
    alignment = Alignment(horizontal="center", vertical="center")

    digit_font = Font(size=15)

    for num in range(0, 10):
        sheet.merge_cells(start_row=1, end_row=1,
                          start_column=header_start,
                          end_column=header_end)

        selectedCell = sheet.cell(1, header_start)
        selectedCell.font = header_font
        selectedCell.alignment = alignment
        selectedCell.fill = header_fill
        selectedCell.value = num

        digit_fill = PatternFill(
            fill_type="solid", fgColor=str(randint(100000, 999999)))

        for rowCount, result in enumerate(results, start=2):
            for colCount, digit in enumerate(result):
                sheet.cell(
                    rowCount, colCount + header_start).font = digit_font
                sheet.cell(
                    rowCount, colCount + header_start).alignment = alignment
                sheet.cell(
                    rowCount, colCount + header_start).value = int(digit)

                if int(digit) == num:
                    sheet.cell(
                        rowCount, colCount + header_start).fill = digit_fill

        header_start = header_end + 2
        header_end = header_end + header_gap + 1

    wb.save("filtered_results.xlsx")


def main():
    export_to_excel()


if __name__ == "__main__":
    main()
