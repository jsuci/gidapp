from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from itertools import islice
from random import randint


def get_all_results():
    """
    INPUT:
        results_v1.1.txt - a text file containing all the results in vertical
        format

    OUTPUT:
        output - a list of results (ex. ["123", "456"...]) taken from
        results_v1.1.txt
    """

    output = []

    with open("results_v1.txt", "r") as fi:
        for entry in islice(fi, 2, None):
            output.insert(0, entry.strip())

    return output


def get_gap_results():
    """
    INPUT:
        all_results - a list of results (ex. ["427", "691", ...]) taken from
        get_all_results()

    OUTPUT:
        output - a dictionary composed of numbers(0-9) as keys and values
        consists of list of tuples containing the gap value and marked results
        (ex. {0: [(1, ["123", ("456", True)])]})

    """

    all_results = get_all_results()
    num = "0"
    step = 1
    gap = step
    found_count = 0

    for count, result in enumerate(all_results):
        if step == count:
            if num in result:
                print(f"{result}*")

                found_count += 1
                step += (gap + 1)

            else:
                break

        print(f"{result}")


def filter_gap_excel():
    results = get_gap_results()[-100:]

    wb = Workbook()
    sheet = wb.active

    header_gap = 3
    header_start = 1
    header_end = 3

    header_font = Font(size=15, bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="33ccff")
    alignment = Alignment(horizontal="center", vertical="center")

    digit_font = Font(size=15)

    for num in range(0, 10):
        sheet.merge_cells(start_row=1, end_row=1,
                          start_column=header_start,
                          end_column=header_end)

        selectedCell = sheet.cell(1, header_start)
        selectedCell.font = header_font
        selectedCell.alignment = alignment
        selectedCell.fill = header_fill
        selectedCell.value = num

        digit_fill = PatternFill(
            fill_type="solid", fgColor=str(randint(696969, 778899)))

        for rowCount, result in enumerate(results, start=2):
            for colCount, digit in enumerate(result):
                sheet.cell(
                    rowCount, colCount + header_start).font = digit_font
                sheet.cell(
                    rowCount, colCount + header_start).alignment = alignment
                sheet.cell(
                    rowCount, colCount + header_start).value = int(digit)

                # if str(num) in result:
                #     sheet.cell(
                #         rowCount, colCount + header_start).fill = digit_fill

                if num == int(digit):
                    sheet.cell(
                        rowCount, colCount + header_start).fill = digit_fill

        header_start = header_end + 2
        header_end = header_end + header_gap + 1

    wb.save("results_filter_excel.xlsx")

    print("Results are up to date")


def main():
    filter_gap_excel()


if __name__ == "__main__":
    # main()
    get_gap_results()
