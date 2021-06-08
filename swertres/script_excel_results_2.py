from openpyxl import Workbook
from openpyxl.styles import (
    NamedStyle, Alignment, Font,
    PatternFill, Border, Side
)
from itertools import islice
from openpyxl.utils import get_column_letter


def check_num(user_num, my_num):
    fdigits = []

    for num in user_num:
        if num in my_num:
            fdigits.append(num)
            my_num = my_num.replace(num, '', 1)

    if len(fdigits) >= 1:
        return True
    else:
        return False


def get_results():
    results = []

    with open("results_v2.txt", "r") as fi:
        for line in islice(fi, 2, None):
            l_list = line.strip().split()
            day = l_list[1]
            date = l_list[0]
            nums = l_list[4:]
            results.append((date, day, nums))

    return results


def write_to_excel():

    wb = Workbook()
    ws = wb.active

    # Set default styles

    bd = Side(style='thin', color="000000")
    bd_thick = Side(style='thick', color="0c1e5f")

    res_style = NamedStyle(name="res_style")
    res_style.font = Font(bold=False, size=21, color='000000')
    res_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    res_style.alignment = Alignment(horizontal="center", vertical="center")
    res_style.fill = PatternFill(start_color="82C09A", fill_type="solid")

    res_macth_style = NamedStyle(name="res_macth_style")
    res_macth_style.font = Font(bold=False, size=21)
    res_macth_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    res_macth_style.alignment = Alignment(
        horizontal="center", vertical="center")
    res_macth_style.fill = PatternFill(start_color="FFFFFF", fill_type="solid")

    def_style = NamedStyle(name="def_style")
    def_style.font = Font(bold=False, size=21)
    def_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    def_style.alignment = Alignment(horizontal="center", vertical="center")

    del wb["Sheet"]

    days = ['sun', 'mon', 'tue', 'wed', 'thu', 'fri', 'sat']

    for i in range(10):
        user_num = str(i)
        results = get_results()

        wb.create_sheet(user_num)
        ws = wb[user_num]

        # day entry
        for i in range(len(days)):

            ws.column_dimensions[get_column_letter(i + 1)].width = 12

            day_cell = ws.cell(
                row=1,
                column=i + 1,
                value=f'{days[i]}'.upper()
            )

            day_cell.style = def_style

        # results entry
        row_loc = 2
        for e in results:
            day = e[1]
            res = e[-1]

            if day == 'sat':
                col_loc = 7

            if day == 'sun':
                col_loc = 1

            if day == 'mon':
                col_loc = 2

            if day == 'tue':
                col_loc = 3

            if day == 'wed':
                col_loc = 4

            if day == 'thu':
                col_loc = 5

            if day == 'fri':
                col_loc = 6

            for rc in range(len(res)):

                res_cell = ws.cell(
                    row=row_loc + rc,
                    column=col_loc,
                    value=f'{res[rc]}'
                )

                if check_num(user_num, res[rc]):
                    res_cell.style = res_macth_style
                else:
                    res_cell.style = res_style

            if day == 'sat':
                for e in ws[row_loc + 2]:
                    e.border = Border(
                        left=bd, top=bd, right=bd, bottom=bd_thick)

                row_loc += 3

    wb.save("excel_results_2.xlsx")


def main():

    write_to_excel()


if __name__ == "__main__":
    main()
