from itertools import islice
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    NamedStyle,
    Border,
    Side
)


def get_results():
    results = []

    with open("results_v1.txt", "r") as fi:
        for entry in islice(fi, 2, None):
            result = entry.strip()
            results.append(result)

    return results


def check_num(user_num, my_num):
    fdigits = []

    for num in user_num:
        if num in my_num:
            fdigits.append(num)
            my_num = my_num.replace(num, '', 1)

    return (''.join(fdigits), len(fdigits))


def duplicate_pairs():
    all_results = get_results()

    output = []

    for count in range(len(all_results) - 1):
        orig_result = all_results[count]
        compared_result = all_results[count + 1]
        matches = check_num(orig_result, compared_result)

        if matches[1] == 2:
            output.append([all_results[count], matches[0]])
        else:
            output.append([all_results[count], ""])

    return output


def export_to_excel(results):
    wb = Workbook()
    sheet = wb.active

    # Set default styles
    cust_style = NamedStyle(name="cust_style")
    cust_style.font = Font(bold=True, size=18)
    bd = Side(style='thin', color="000000")
    cust_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    cust_style.alignment = Alignment(horizontal="center", vertical="center")

    for row_count, each_result in enumerate(results, start=1):

        for column_count, each_digit in enumerate(each_result[0], start=1):
            sheet.cell(
                row_count,
                column_count
            ).style = cust_style

            sheet.cell(
                row_count,
                column_count
            ).value = each_digit

        if each_result[1]:
            sheet.cell(row_count, 6).style = cust_style
            sheet.cell(row_count, 6).value = each_result[1]

    wb.save("duplicate_pairs.xlsx")


def main():

    print("Generating duplicate pairs, please wait...")

    results = duplicate_pairs()

    export_to_excel(results)

    print("Done exporting.")


if __name__ == "__main__":
    main()
