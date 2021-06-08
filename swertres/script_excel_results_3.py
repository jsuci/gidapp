from re import split
from sys import argv
from itertools import islice
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    NamedStyle,
    Border,
    Side
)


def get_results(gap):
    results = []
    reversed_res = []

    with open("results_v2.txt", "r") as fi:
        for entry in islice(fi, 2, None):
            result = split(r"\s{2,}", entry.strip())
            reversed_res.insert(0, result)

    for c, e in enumerate(reversed_res):
        if c % gap == 0:
            e.append('t')
        else:
            e.append('f')

        results.insert(0, e)

    return results


def check_num(user_num, my_num):
    fdigits = []

    for num in user_num:
        if num in my_num:
            fdigits.append(num)
            my_num = my_num.replace(num, '', 1)

    return (''.join(fdigits), len(fdigits))


def export_to_excel(results, gap):
    wb = Workbook()
    sheet = wb.active

    # Set default styles
    cust_style = NamedStyle(name="cust_style")
    cust_style.font = Font(bold=True, size=18)
    bd = Side(style='thin', color="000000")
    cust_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    cust_style.alignment = Alignment(horizontal="center", vertical="center")

    select_style = NamedStyle(name="select_style")
    select_style.font = Font(bold=True, size=18)
    bd = Side(style='thin', color="000000")
    select_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    select_style.alignment = Alignment(horizontal="center", vertical="center")
    select_style.fill = PatternFill(start_color="FFC7CE", fill_type="solid")

    sheet.column_dimensions['A'].width = 33

    for row_count, date_results in enumerate(results, start=1):

        sheet.cell(row_count, 1).style = cust_style
        sheet.cell(row_count, 1).alignment = Alignment(
            horizontal="right")
        sheet.cell(row_count, 1).value = date_results[0].upper()

        if date_results[-1] == 't':
            sheet.cell(row_count, 3 + 2).style = select_style
            sheet.cell(row_count, 4 + 2).style = select_style
            sheet.cell(row_count, 5 + 2).style = select_style
        else:
            sheet.cell(row_count, 3 + 2).style = cust_style
            sheet.cell(row_count, 4 + 2).style = cust_style
            sheet.cell(row_count, 5 + 2).style = cust_style

            sheet.cell(
                row_count, 3 + 2).fill = PatternFill(
                    start_color="b4c9ea", fill_type="solid")

            sheet.cell(
                row_count, 4 + 2).fill = PatternFill(
                    start_color="b4c9ea", fill_type="solid")

            sheet.cell(
                row_count, 5 + 2).fill = PatternFill(
                    start_color="b4c9ea", fill_type="solid")

        sheet.cell(row_count, 3 + 2).value = date_results[1]
        sheet.cell(row_count, 4 + 2).value = date_results[2]
        sheet.cell(row_count, 5 + 2).value = date_results[3]

    wb.save("excel_results_3.xlsx")


def main():

    print("Generating excel 3 results, please wait...")

    if len(argv) < 2:
        gap = 2
    else:
        gap = int(argv[1])

    results = get_results(gap)

    export_to_excel(results, gap)

    print("Done exporting.")


if __name__ == "__main__":
    main()
