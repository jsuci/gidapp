from itertools import islice
from re import split
from sys import argv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    NamedStyle,
    Border,
    Side
)


def get_results(user_filter):
    output = {}

    with open("results_v2.txt", "r") as fi:

        for entry in islice(fi, 2, None):
            entry = split(r"\s{2,}", entry.strip())

            if user_filter[-1] == '-s':
                pass
                res = strict_filter(entry[1:], user_filter[:-1])
            else:
                res = loose_filter(entry[1:], user_filter[:-1])

            date = split(r"\s{1,}", entry[0].strip())

            output.setdefault(date[3], {})
            output[date[3]].setdefault(date[2], [])
            output[date[3]][date[2]].append((date[0], res))

        return output


def check_num(user_num, my_num):
    fdigits = []

    for num in user_num:
        if num in my_num:
            fdigits.append(num)
            my_num = my_num.replace(num, '', 1)

    return (''.join(fdigits), len(fdigits))


def strict_filter(entry_num, comp_num):
    output = []

    for e in entry_num:

        # prevent appending same number to output
        # detect if there is a match
        # since we want to return same entry_num length
        check_count = 0

        for c in comp_num:
            if len(c) == 2 and e[:2] == c:
                check_count += 1

            if len(c) == 3 and e == c:
                check_count += 1

        if check_count > 0:
            output.append((e, 't'))
        else:
            output.append((e, 'f'))

    return output


def loose_filter(entry_num, comp_num):

    output = []

    for e in entry_num:
        check_count = 0

        for c in comp_num:
            check = check_num(e, c)

            if len(comp_num[0]) == check[1]:
                check_count += 1

        if check_count > 0:
            output.append((e, 't'))
        else:
            output.append((e, 'f'))

    return output


def export_to_excel(results):
    wb = Workbook()
    ws = wb.active

    # Set default styles
    res_style = NamedStyle(name="res_style")
    res_style.font = Font(bold=True, size=21, color='000000')
    bd = Side(style='thin', color="000000")
    res_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    res_style.alignment = Alignment(horizontal="center", vertical="center")
    res_style.fill = PatternFill(start_color="82C09A", fill_type="solid")

    res_macth_style = NamedStyle(name="res_macth_style")
    res_macth_style.font = Font(bold=True, size=21)
    bd = Side(style='thin', color="000000")
    res_macth_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    res_macth_style.alignment = Alignment(
        horizontal="center", vertical="center")
    res_macth_style.fill = PatternFill(start_color="FFFFFF", fill_type="solid")

    def_style = NamedStyle(name="def_style")
    def_style.font = Font(bold=True, size=21)
    bd = Side(style='thin', color="000000")
    def_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    def_style.alignment = Alignment(horizontal="center", vertical="center")

    time = ['2PM', '4PM', '9PM']
    col_colors = [
        "ff6699", "ff6666", "ff6633", "ccff33",
        "9966ff", "66cc33", "6699ff", "6666ff",
        "33ff33", "339933", "3333ff", "009966"
    ]

    # year entry
    date_loc = 1
    for year, month in sorted(results.items()):

        # date entry
        date_cell = ws.cell(
            row=date_loc,
            column=1,
            value="Date".upper()
        )
        date_cell.style = def_style

        date_merge = ws.merge_cells(
            start_row=date_loc,
            end_row=date_loc + 1,
            start_column=1,
            end_column=1
        )

        for i in range(1, 32):
            date_num_cell = ws.cell(
                row=i + date_loc + 1,
                column=1,
                value=f"{i:0>2}"
            )
            date_num_cell.style = def_style

        # month year entry
        mo_yr_loc = 1
        color_count = 0
        for month, date_res in month.items():

            ws.column_dimensions[get_column_letter(1)].width = 12
            ws.column_dimensions[get_column_letter(mo_yr_loc + 1)].width = 9
            ws.column_dimensions[get_column_letter(mo_yr_loc + 2)].width = 9
            ws.column_dimensions[get_column_letter(mo_yr_loc + 3)].width = 9

            mo_yr_cell = ws.cell(
                row=date_loc,
                column=mo_yr_loc + 1,
                value=f"{month} {year}".upper()
            )

            mo_yr_cell.style = def_style

            mo_yr_merge = ws.merge_cells(
                start_row=date_loc,
                end_row=date_loc,
                start_column=mo_yr_loc + 1,
                end_column=mo_yr_loc + 3
            )

            # time entry
            time_cell_2pm = ws.cell(
                row=date_loc + 1,
                column=mo_yr_loc + 1,
                value=f"{time[0]}".upper()
            )
            time_cell_2pm.style = def_style

            time_cell_4pm = ws.cell(
                row=date_loc + 1,
                column=mo_yr_loc + 2,
                value=f"{time[1]}".upper()
            )
            time_cell_4pm.style = def_style

            time_cell_9pm = ws.cell(
                row=date_loc + 1,
                column=mo_yr_loc + 3,
                value=f"{time[2]}".upper()
            )
            time_cell_9pm.style = def_style

            # results entry
            for res_row in range(31):
                try:
                    res = date_res[res_row][-1]
                    for rcount, r in enumerate(res):
                        if r[1] == 't':
                            res_cell = ws.cell(
                                row=res_row + 1 + date_loc + 1,
                                column=rcount + mo_yr_loc + 1,
                                value=r[0]
                            )
                            res_cell.style = res_macth_style

                        if r[1] == 'f':
                            res_cell = ws.cell(
                                row=res_row + 1 + date_loc + 1,
                                column=rcount + mo_yr_loc + 1,
                                value=r[0]
                            )
                            res_cell.style = res_style
                            res_cell.fill = PatternFill(
                                fill_type="solid",
                                fgColor=col_colors[color_count]
                            )
                except Exception:
                    for i in range(3):
                        res_cell = ws.cell(
                            row=res_row + 1 + date_loc + 1,
                            column=i + mo_yr_loc + 1,
                            value='-'
                        )
                        res_cell.style = res_style
                        res_cell.fill = PatternFill(
                            fill_type="solid",
                            fgColor=col_colors[color_count]
                        )

            color_count += 1
            mo_yr_loc += 3
        date_loc += 33

    wb.save("excel_results_4.xlsx")


def main():

    results = get_results(argv[1:])

    print("Generating excel results 4 file, please wait...")

    export_to_excel(results)

    print("Done exporting.")


if __name__ == "__main__":
    main()
