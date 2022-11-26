from itertools import islice
from re import split
from openpyxl import Workbook
from openpyxl.styles import (
    NamedStyle, Alignment, Font,
    PatternFill, Border, Side
)
from datetime import datetime
from sys import argv


def read_results():
    output = []

    with open("results_v2.txt", "r") as fi:
        # updated: 04 sat jan 2020 2
        check_date = fi.readline().strip()[-1]

        for entry in islice(fi, 1, None):
            entry = split(r"\s{2,}", entry.strip())
            entry[0] = datetime.strptime(
                entry[0],
                "%d %a %b %Y"
            ).strftime(
                "%d %B %Y"
            )

            output.append(entry)

        return output
        # if check_date == "2":
        #     return output
        # else:
        #     return output[:-1]


def fix_results():
    all_results = read_results()
    month_results = []
    temp_results = []
    output = []

    for i in range(1, len(all_results)):
        # pev_month = January
        prev_month = all_results[i - 1][0].split(" ")[1]
        next_month = all_results[i][0].split(" ")[1]

        if prev_month == next_month:
            temp_results.append(all_results[i - 1])
        else:
            temp_results.append(all_results[i - 1])
            month_results.append(temp_results)
            temp_results = []

    # Append the last entry
    temp_results.append(all_results[-1])
    month_results.append(temp_results)

    for month_res in month_results:
        for i in range(32 - 1):
            date_one = int(f"{i + 1:02d}")
            add_entry = [
                f"{date_one:02d} {month_res[0][0][3:]}",
                "-",
                "-",
                "-"
            ]

            try:
                month_res_index = month_res.index(month_res[i])
                date_two = int(month_res[i][0][:2])
            except Exception:
                month_res.append(add_entry)
                continue

            if date_one != date_two:
                date_one_int = i
                date_two_int = int(month_res[i][0][:2])

                if date_one_int < date_two_int:
                    month_res.insert(month_res_index, add_entry)

        output.append(month_res)

    return output


def excel_export(user_res, user_option):
    def exact_match(res, user_res):
        if res == user_res:
            return True
        else:
            return False

    def check_num(user_num, my_num):
        count = 0

        for num in user_num:
            if num in my_num:
                count += 1
                my_num = my_num.replace(num, '', 1)

        if count == 1:
            return "one"
        elif count == 2:
            return "two"
        elif count == 3:
            return "three"
        else:
            return "zero"

    # Creating Workbook
    wb = Workbook()
    ws = wb.active

    # Formatting
    col_colors = [
        "ff6699", "ff6666", "ff6633", "ccff33",
        "9966ff", "66cc33", "6699ff", "6666ff",
        "33ff33", "339933", "3333ff", "009966"
    ]
    bd = Side(style="thin", color="555555")
    cell_style = NamedStyle(
        name="cell_style",
        font=Font(size=15, bold=False),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=Border(top=bd, bottom=bd, right=bd, left=bd)
    )
    wb.add_named_style(cell_style)

    # Control group
    month_results = fix_results()
    month_gap = 1
    year_gap = 1
    color_count = 0
    time_str = ["11AM", "4PM", "9PM"]

    for month_res in month_results:

        # Control group for finding the combis
        gap_match_count = 0
        match_count = 0
        found_index = 0
        first_match_index = 0
        second_match_index = 0
        before_match = []
        after_match = []

        # res = ['24 September 2019', '091', '321', '729']
        for c_res, res in enumerate(month_res):

            # res_date_month = ['31', 'December', '2017']
            res_date_month = res[0].split(" ")
            res_day = res_date_month[0]
            res_month = res_date_month[1]
            res_year = res_date_month[2]
            digits = res[1:]

            # Date header (A1:A2)
            ws.merge_cells(
                start_row=year_gap,
                end_row=year_gap + 1,
                start_column=1,
                end_column=1
            )

            date_cell = ws.cell(row=year_gap, column=1, value="Date")
            date_cell.style = cell_style

            # Date entry
            ws.cell(
                row=year_gap + c_res + 2,
                column=1,
                value=res_day
            ).style = cell_style

            # Month header
            ws.merge_cells(
                start_row=year_gap,
                end_row=year_gap,
                start_column=month_gap + 1,
                end_column=month_gap + 3
            )

            month_cell = ws.cell(
                row=year_gap,
                column=month_gap + 1,
                value=f"{res_month} {res_year}"
            )

            month_cell.style = cell_style

            # Time header
            for c_time, time in enumerate(time_str, start=1):
                time_cell = ws.cell(
                    row=year_gap + 1,
                    column=month_gap + c_time,
                    value=time
                )
                time_cell.style = cell_style

            # Result entry
            for dg_c, dg in enumerate(digits):

                # row - controls the top and bottom movement
                # column - controls the left and right movement
                result_cell = ws.cell(
                    row=c_res + year_gap + 2,
                    column=dg_c + month_gap + 1,
                    value=dg
                )

                result_cell.style = cell_style
                result_cell.fill = PatternFill(
                    fill_type="solid", fgColor=col_colors[color_count]
                )

                for u_res in user_res:
                    if user_option == "y" and exact_match(dg, u_res):
                        result_cell.fill = PatternFill(
                            fill_type="solid",
                            fgColor="FFFFFF"
                        )
                        # set found_index to the index of the current match
                        # increment match_count by 1 for every match
                        found_index = month_res.index(res)
                        match_count += 1

                    elif len(u_res) == 3:
                        if (user_option == "n"
                                and check_num(dg, u_res) == 'three'):

                            result_cell.fill = PatternFill(
                                fill_type="solid",
                                fgColor="FFFFFF"
                            )
                    elif len(u_res) == 2:
                        if (user_option == "n"
                                and check_num(dg, u_res) == 'two'):

                            result_cell.fill = PatternFill(
                                fill_type="solid",
                                fgColor="FFFFFF"
                            )

            # match_count 1 find before_match value
            # match_count 2 find after_match value
            if match_count == 1:
                first_match_index = found_index

            if match_count == 2:
                second_match_index = found_index

        gap_match_count = second_match_index - first_match_index

        if gap_match_count >= 1:

            if (first_match_index - gap_match_count) >= 0:
                before_match = month_res[first_match_index - gap_match_count]

                if "-" not in before_match:
                    print(f"{before_match}")

            if (second_match_index + gap_match_count) < len(month_res):
                after_match = month_res[second_match_index + gap_match_count]

                if "-" not in after_match:
                    print(f"{after_match}")

        if res_day == "31" and res_month == "December":
            year_gap += 35
            month_gap = 1
            color_count = 0
        else:
            month_gap += 3
            color_count += 1

    wb.save("excel_results.xlsx")


def main():
    user_res = argv

    if len(user_res) < 2:
        user_option = ""
    else:
        user_option = input("Exact match? (y/n): ")

    print("Generating excel file, please wait...")

    excel_export(user_res[1:], user_option)

    print("Done exporting.")


if __name__ == "__main__":
    main()
