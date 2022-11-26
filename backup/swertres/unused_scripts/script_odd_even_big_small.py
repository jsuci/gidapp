from itertools import islice
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    NamedStyle,
    Border,
    Side
)


def get_results():
    results = []

    with open("results_v1.txt", "r") as fi:
        for entry in islice(fi, 2, None):
            result = entry.strip()
            results.append(result)

    return results


def odd_even_big_small():
    all_results = get_results()

    odd = ['1', '3', '5', '7', '9']
    even = ['2', '4', '6', '8', '0']
    big = ['6', '7', '8', '9', '0']
    small = ['1', '2', '3', '4', '5']

    output = []

    for each_result in all_results:
        per_result = []
        per_result.append(each_result)

        for each_digit in each_result:
            per_digit = []

            if each_digit in big:
                per_digit.append('B')

            if each_digit in small:
                per_digit.append('S')

            if each_digit in odd:
                per_digit.append('O')

            if each_digit in even:
                per_digit.append('E')

            per_result.append("".join(per_digit))

        output.append(per_result)

    return output


def export_to_excel(results):
    wb = Workbook()
    sheet = wb.active

    # Set default styles
    cust_style = NamedStyle(name="cust_style")
    cust_style.font = Font(bold=True, size=18)
    bd = Side(style='thin', color="000000")
    cust_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    cust_style.alignment = Alignment(horizontal="center", vertical="center")

    for row_count, each_result in enumerate(results, start=1):

        for column_count, each_digit in enumerate(each_result[0], start=1):
            sheet.cell(
                row_count,
                column_count
            ).style = cust_style

            sheet.cell(
                row_count,
                column_count
            ).value = each_digit

        sheet.cell(row_count, 6).style = cust_style
        sheet.cell(row_count, 6).value = each_result[1]

        sheet.cell(row_count, 7).style = cust_style
        sheet.cell(row_count, 7).value = each_result[2]

        sheet.cell(row_count, 8).style = cust_style
        sheet.cell(row_count, 8).value = each_result[3]

    wb.save("odd_even_big_small.xlsx")


def main():

    print("Generating odd even big small file, please wait...")

    results = odd_even_big_small()
    export_to_excel(results)

    print("Done exporting.")


if __name__ == "__main__":
    main()
